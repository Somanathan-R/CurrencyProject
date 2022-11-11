import openpyxl as xl
from forex_python.converter import CurrencyCodes,CurrencyRates

wb=xl.load_workbook('RawData.xlsx')
sheet = wb['Sheet1']


Rates= CurrencyRates()

for row in range(2,sheet.max_row+1):
    code = sheet.cell(row,3).value
    amount = sheet.cell(row,2).value
    dollar = sheet.cell(row,4)
    dollar.value = Rates.convert(code,'USD',amount)


wb.save('CurrencyNew.xlsx')