import openpyxl as xl
wb= xl.load_workbook('try.xlsx')
sheet= wb['Sheet1']
#cell = sheet['a1']

for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    new_price = cell.value*0.9
    new_price_cell = sheet.cell(row,4)
    new_price_cell.value = new_price


wb.save('try2.xlsx')